import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os

def export_report(df,summary,charts):
    report_dir = "report"
    if not os.path.exists(report_dir):
        os.makedirs(report_dir)

    output_file = os.path.join(report_dir,"销售报表.xlsx")

    with pd.ExcelWriter(output_file) as writer:
        df.to_excel(writer,sheet_name="清洗数据",index=False)
        summary.to_excel(writer,sheet_name="销售汇总")

    wb = load_workbook(output_file)
    ws = wb.create_sheet("图表")

    positions = ["A1","A20","A40"]
    for i,chart in enumerate(charts):
        img = Image(chart)
        ws.add_image(img,positions[i])
    wb.save(output_file)

    return output_file

